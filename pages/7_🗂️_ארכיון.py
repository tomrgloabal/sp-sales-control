"""
ארכיון מאוחד — חיפוש חופשי לפי כל עמודה, סטטוס פקדון חי/נגמר
"""
import json
import openpyxl
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import date
from auth import require_login, current_user
from sheets import read_df, append_row, log_action

require_login()

st.markdown("""<style>
.stApp{direction:rtl;}
header[data-testid="stHeader"]{display:none;}
.status-badge {
    display:inline-block; border-radius:6px; padding:2px 10px;
    font-size:.82rem; font-weight:600;
}
</style>""", unsafe_allow_html=True)
st.markdown("<h2 style='color:#1E2761;'>🗂️ ארכיון מאוחד</h2>", unsafe_allow_html=True)

DATA_DIR      = Path(__file__).parent.parent / "local_data"
INVESTORS_FILE = DATA_DIR / "ProductInvestors.json"
EXCEL_DIRS    = [Path(__file__).parent.parent,
                 Path(__file__).parent.parent.parent / "SP-SALES"]

# ── Import Excel ──────────────────────────────────────────────────────────────
def _find_excel() -> Path | None:
    for d in EXCEL_DIRS:
        if d.exists():
            for f in d.iterdir():
                if f.suffix == ".xlsx" and "גיוס" in f.name:
                    return f
    return None

def _parse_excel(xlsx_path: Path) -> dict:
    wb  = openpyxl.load_workbook(xlsx_path)
    ws  = wb["גיוסים"] if "גיוסים" in wb.sheetnames else wb[wb.sheetnames[0]]
    result: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        isin = row[6]
        if not isin or not str(isin).strip().startswith("XS"):
            continue
        isin          = str(isin).strip()
        investor_name = row[1]
        product_name  = row[2]
        status        = row[3]
        partner       = row[4]
        bank          = row[5]
        raw_amount    = row[7]
        currency      = row[8]
        issue_date    = row[10]
        issuer        = row[12]
        try:
            amount = float(raw_amount) if raw_amount and str(raw_amount).replace(".", "").replace("-", "").isdigit() else 0
        except Exception:
            amount = 0
        if isin not in result:
            result[isin] = {
                "שם מלא":     str(product_name).strip() if product_name else "",
                "מנפיק":      str(issuer).strip()        if issuer        else "",
                "ISSUE DATE": issue_date.strftime("%Y-%m-%d") if hasattr(issue_date, "strftime") else str(issue_date or ""),
                "משקיעים":    [],
            }
        if investor_name:
            result[isin]["משקיעים"].append({
                "שם המשקיע": str(investor_name).strip(),
                "סכום":      amount,
                "מטבע":      str(currency or "ILS"),
                "שותף":      str(partner or "").strip(),
                "בנק":       str(bank or "").strip(),
                "סטטוס":     str(status or "").strip(),
            })
    return result

with st.expander("🔄 ייבוא / עדכון ארכיון מ-Excel", expanded=False):
    xlsx_path = _find_excel()
    if xlsx_path:
        st.success(f"נמצא קובץ: `{xlsx_path.name}`")
        if st.button("📥 טען / עדכן ארכיון", use_container_width=True, type="primary"):
            data = _parse_excel(xlsx_path)
            INVESTORS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            log_action(current_user(), "טעינת ארכיון מ-Excel", f"{len(data)} ISINs")
            st.success(f"✓ נטענו {len(data)} ISINs")
            st.cache_data.clear()
            st.rerun()
    else:
        st.warning("לא נמצא קובץ Excel עם 'גיוס' בשם. ניתן להעלות ידנית:")
        uploaded = st.file_uploader("העלה קובץ Excel", type=["xlsx"])
        if uploaded:
            tmp = DATA_DIR / "uploaded_investors.xlsx"
            tmp.write_bytes(uploaded.read())
            data = _parse_excel(tmp)
            INVESTORS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            log_action(current_user(), "טעינת ארכיון - ידנית", f"{len(data)} ISINs")
            st.success(f"✓ נטענו {len(data)} ISINs")
            st.cache_data.clear()
            st.rerun()

st.divider()

# ── Build unified flat DataFrame ──────────────────────────────────────────────
@st.cache_data(ttl=90)
def _build_df() -> pd.DataFrame:
    if not INVESTORS_FILE.exists():
        return pd.DataFrame()
    archive = json.loads(INVESTORS_FILE.read_text(encoding="utf-8"))

    # Flatten: one row per investor per ISIN
    rows = []
    for isin, prod in archive.items():
        for inv in prod.get("משקיעים", []):
            rows.append({
                "ISIN":         isin,
                "שם מוצר":      prod.get("שם מלא", "")[:70],
                "מנפיק":        prod.get("מנפיק", ""),
                "תאריך הנפקה":  prod.get("ISSUE DATE", ""),
                "שם משקיע":     inv.get("שם המשקיע", ""),
                "סכום":         inv.get("סכום", 0),
                "מטבע":         inv.get("מטבע", "ILS"),
                "שותף":         inv.get("שותף", ""),
                "בנק":          inv.get("בנק", ""),
                "סטטוס משקיע":  inv.get("סטטוס", ""),
            })
    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # Cross-reference Products sheet → deposit status
    products_df = read_df("Products")
    status_map: dict[str, str] = {}
    if not products_df.empty and "ISIN" in products_df.columns:
        for _, r in products_df.iterrows():
            k = str(r.get("ISIN", "")).strip()
            if k:
                status_map[k] = str(r.get("סטטוס", ""))

    # Cross-reference Redemptions
    red_df = read_df("Redemptions")
    redeemed: set[str] = set()
    if not red_df.empty and "ISIN" in red_df.columns:
        redeemed = set(red_df["ISIN"].dropna().str.strip())

    def _dep_status(isin: str) -> str:
        if isin in status_map:
            s = status_map[isin]
            if s == "פעיל":  return "🟢 פעיל"
            if s == "סגור":  return "🔴 סגור"
            return f"📋 {s}"
        if isin in redeemed:
            return "⚠️ פקע"
        return "📁 ארכיון"

    df["סטטוס פקדון"] = df["ISIN"].apply(_dep_status)
    return df

df = _build_df()

if df.empty:
    st.info("הארכיון ריק. ייבא קובץ Excel למעלה.")
    st.stop()

# ── KPIs ──────────────────────────────────────────────────────────────────────
k1, k2, k3, k4 = st.columns(4)
active_isins = df[df["סטטוס פקדון"] == "🟢 פעיל"]["ISIN"].nunique()
closed_isins = df[df["סטטוס פקדון"].isin(["🔴 סגור","⚠️ פקע","📁 ארכיון"])]["ISIN"].nunique()
with k1: st.metric("ISINs בארכיון",   df["ISIN"].nunique())
with k2: st.metric("משקיעים ייחודיים", df["שם משקיע"].nunique())
with k3: st.metric("פקדונות פעילים",   active_isins)
with k4: st.metric("פקדונות שנסגרו",   closed_isins)

st.divider()

# ── Search UI ─────────────────────────────────────────────────────────────────
SEARCHABLE_COLS = ["הכל"] + [
    "ISIN", "שם מוצר", "מנפיק", "שם משקיע", "שותף", "בנק", "מטבע", "סטטוס פקדון", "תאריך הנפקה"
]

sc1, sc2, sc3 = st.columns([4, 2, 2])
with sc1:
    query = st.text_input("🔍 חיפוש חופשי", placeholder="שם משקיע / ISIN / מנפיק / בנק...", label_visibility="collapsed")
with sc2:
    search_col = st.selectbox("חפש לפי עמודה", SEARCHABLE_COLS, label_visibility="collapsed")
with sc3:
    status_filter = st.selectbox("סטטוס פקדון", ["הכל", "🟢 פעיל", "🔴 סגור", "⚠️ פקע", "📁 ארכיון"], label_visibility="collapsed")

# Apply filters
filtered = df.copy()

if status_filter != "הכל":
    filtered = filtered[filtered["סטטוס פקדון"] == status_filter]

if query:
    q = query.strip()
    if search_col == "הכל":
        # Search across all text columns
        text_cols = ["ISIN", "שם מוצר", "מנפיק", "שם משקיע", "שותף", "בנק", "מטבע", "סטטוס פקדון", "תאריך הנפקה"]
        mask = pd.Series(False, index=filtered.index)
        for col in text_cols:
            if col in filtered.columns:
                mask |= filtered[col].astype(str).str.contains(q, case=False, na=False)
        filtered = filtered[mask]
    else:
        if search_col in filtered.columns:
            filtered = filtered[filtered[search_col].astype(str).str.contains(q, case=False, na=False)]

# ── Results table ─────────────────────────────────────────────────────────────
st.markdown(f"**{len(filtered):,} תוצאות**")

# Display columns — put סטטוס פקדון first for visibility
display_cols = ["סטטוס פקדון", "ISIN", "שם מוצר", "מנפיק", "תאריך הנפקה",
                "שם משקיע", "סכום", "מטבע", "שותף", "בנק", "סטטוס משקיע"]
display_cols = [c for c in display_cols if c in filtered.columns]

def _color_status(val):
    colors = {
        "🟢 פעיל":   "background-color:#C6EFCE; color:#1A7A4A",
        "🔴 סגור":   "background-color:#FDEBD0; color:#C55A11",
        "⚠️ פקע":    "background-color:#FFF2CC; color:#7D6608",
        "📁 ארכיון": "background-color:#EEF0F8; color:#555",
    }
    return colors.get(val, "")

styled = (
    filtered[display_cols]
    .style
    .map(_color_status, subset=["סטטוס פקדון"])
    .format({"סכום": lambda x: f"₪{x:,.0f}" if isinstance(x, (int, float)) else x})
)
st.dataframe(styled, use_container_width=True, hide_index=True, height=420)

# ── Export ────────────────────────────────────────────────────────────────────
import io
buf = io.BytesIO()
with pd.ExcelWriter(buf, engine="openpyxl") as writer:
    filtered[display_cols].to_excel(writer, index=False, sheet_name="ארכיון")
st.download_button(
    "📥 ייצוא Excel",
    buf.getvalue(),
    f"archive_search_{date.today()}.xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=False,
)

st.divider()

# ── Actions on selected rows ──────────────────────────────────────────────────
st.markdown("### ⚡ פעולות מהירות")

# Select ISIN for action
isin_options = sorted(filtered["ISIN"].unique().tolist())
if not isin_options:
    st.info("אין תוצאות לפעולה.")
    st.stop()

act1, act2 = st.columns(2)

# ── Add to Pipeline ───────────────────────────────────────────────────────────
with act1:
    st.markdown("**→ הוסף לפייפליין הנוכחי**")
    sel_isin_pipe = st.selectbox("ISIN", isin_options, key="pipe_isin_sel")
    isin_investors = filtered[filtered["ISIN"] == sel_isin_pipe]["שם משקיע"].dropna().unique().tolist()
    sel_inv_pipe = st.multiselect("בחר משקיעים", isin_investors, key="pipe_inv_sel")

    if sel_inv_pipe and st.button("➕ הוסף לפייפליין", use_container_width=True, type="primary", key="pipe_add_btn"):
        existing_pipe = read_df("Pipeline")
        already = set(existing_pipe["שם לקוח"].dropna().tolist()) if not existing_pipe.empty and "שם לקוח" in existing_pipe.columns else set()
        added = 0
        for name in sel_inv_pipe:
            if name in already:
                continue
            inv_row = filtered[(filtered["ISIN"] == sel_isin_pipe) & (filtered["שם משקיע"] == name)].iloc[0]
            pipeline_row = [
                name, "", inv_row.get("שותף", current_user()), "A",
                sel_isin_pipe, int(inv_row.get("סכום", 0)), inv_row.get("מטבע", "ILS"),
                "בינונית", date.today().strftime("%d/%m/%Y"), "לא פנו",
                f"ארכיון — {inv_row.get('שם מוצר','')[:40]} | בנק: {inv_row.get('בנק','')}",
                date.today().strftime("%d/%m/%Y"),
            ]
            append_row("Pipeline", pipeline_row)
            added += 1
        log_action(current_user(), "הוספה לפייפליין מארכיון", f"{sel_isin_pipe} | {added} משקיעים")
        st.success(f"✓ {added} משקיעים נוספו לפייפליין")
        if added < len(sel_inv_pipe):
            st.caption(f"ℹ️ {len(sel_inv_pipe)-added} כבר קיימים בפייפליין")

# ── Add to Redemptions ────────────────────────────────────────────────────────
with act2:
    st.markdown("**→ הוסף לפקדונות פקועים**")
    sel_isin_red = st.selectbox("ISIN", isin_options, key="red_isin_sel")
    isin_investors_r = filtered[filtered["ISIN"] == sel_isin_red]["שם משקיע"].dropna().unique().tolist()
    sel_inv_red = st.multiselect("בחר משקיעים", isin_investors_r, key="red_inv_sel")

    if sel_inv_red and st.button("➕ הוסף לפקועים", use_container_width=True, key="red_add_btn"):
        existing_red = read_df("Redemptions")
        already_r = set(existing_red["שם לקוח"].dropna().tolist()) if not existing_red.empty and "שם לקוח" in existing_red.columns else set()
        added_r = 0
        for name in sel_inv_red:
            if name in already_r:
                continue
            inv_row = filtered[(filtered["ISIN"] == sel_isin_red) & (filtered["שם משקיע"] == name)].iloc[0]
            prod_info = filtered[filtered["ISIN"] == sel_isin_red].iloc[0]
            red_row = [
                name, sel_isin_red, prod_info.get("מנפיק", ""),
                int(inv_row.get("סכום", 0)), inv_row.get("מטבע", "ILS"),
                "", "", "לא", "נמוכה",
                f"ארכיון — {prod_info.get('שם מוצר','')[:50]}",
            ]
            append_row("Redemptions", red_row)
            added_r += 1
        log_action(current_user(), "הוספה לפקועים מארכיון", f"{sel_isin_red} | {added_r} משקיעים")
        st.success(f"✓ {added_r} נוספו לפקדונות פקועים")
