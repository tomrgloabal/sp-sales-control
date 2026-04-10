import streamlit as st
import pandas as pd
import json
import os
import openpyxl
from pathlib import Path
from auth import require_login, current_user
from sheets import read_df, append_row, log_action
from config import CONFIDENCE, REDEMP_COLS

require_login()

st.markdown("""<style>.stApp{direction:rtl;} header[data-testid="stHeader"]{display:none;}</style>""", unsafe_allow_html=True)
st.markdown("<h2 style='color:#1E2761;'>🔍 ארכיון פקדונות — חיפוש ISIN</h2>", unsafe_allow_html=True)
st.caption("טען פקדון ישן לפי ISIN — ראה את כל המשקיעים ופרטי המוצר.")

DATA_DIR = Path(__file__).parent.parent / "local_data"
INVESTORS_FILE = DATA_DIR / "ProductInvestors.json"
EXCEL_SEARCH_DIRS = [
    Path(__file__).parent.parent,
    Path(__file__).parent.parent.parent / "SP-SALES",
]


def _find_excel() -> Path | None:
    for d in EXCEL_SEARCH_DIRS:
        if d.exists():
            for f in d.iterdir():
                if f.suffix == ".xlsx" and "גיוס" in f.name:
                    return f
    return None


def _parse_excel(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    sheet_name = "גיוסים"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    result: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        isin = row[6]
        if not isin or not str(isin).strip().startswith("XS"):
            continue
        isin = str(isin).strip()
        investor_name = row[1]
        product_name  = row[2]
        status        = row[3]
        partner       = row[4]
        bank          = row[5]
        raw_amount    = row[7]
        currency      = row[8]
        issue_date    = row[10]
        issuer        = row[12]

        try:
            amount = float(raw_amount) if raw_amount and str(raw_amount).replace(".", "").replace("-", "").isdigit() else 0
        except Exception:
            amount = 0

        if isin not in result:
            result[isin] = {
                "שם מלא":    str(product_name).strip() if product_name else "",
                "מנפיק":     str(issuer).strip()        if issuer        else "",
                "ISSUE DATE": issue_date.strftime("%Y-%m-%d") if hasattr(issue_date, "strftime") else str(issue_date) if issue_date else "",
                "משקיעים":   [],
            }
        if investor_name:
            result[isin]["משקיעים"].append({
                "שם המשקיע": str(investor_name).strip(),
                "סכום":      amount,
                "מטבע":      str(currency) if currency else "ILS",
                "שותף":      str(partner).strip()  if partner  else "",
                "בנק":       str(bank).strip()      if bank      else "",
                "סטטוס":     str(status).strip()    if status    else "",
            })
    return result


def _load_investors() -> dict:
    if INVESTORS_FILE.exists():
        return json.loads(INVESTORS_FILE.read_text(encoding="utf-8"))
    return {}


def _save_investors(data: dict):
    INVESTORS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ── Import / Refresh ──────────────────────────────────────────────────────────
with st.expander("🔄 טעינה / רענון מקובץ Excel", expanded=False):
    xlsx_path = _find_excel()
    if xlsx_path:
        st.success(f"נמצא קובץ: `{xlsx_path.name}`")
        if st.button("📥 טען / עדכן ארכיון מהקובץ", use_container_width=True, type="primary"):
            data = _parse_excel(xlsx_path)
            _save_investors(data)
            log_action(current_user(), "טעינת ארכיון מ-Excel", f"{len(data)} ISINs")
            st.success(f"✓ נטענו {len(data)} ISINs עם משקיעיהם")
            st.rerun()
    else:
        st.warning("לא נמצא קובץ 'גיוס סטראקצרים.xlsx'. הנח אותו בתיקיית sp-sales-app ולחץ רענן.")
        st.caption("ניתן גם להעלות קובץ ידנית:")
        uploaded = st.file_uploader("העלה קובץ Excel", type=["xlsx"])
        if uploaded:
            tmp_path = DATA_DIR / "uploaded_investors.xlsx"
            tmp_path.write_bytes(uploaded.read())
            data = _parse_excel(tmp_path)
            _save_investors(data)
            log_action(current_user(), "טעינת ארכיון - העלאה ידנית", f"{len(data)} ISINs")
            st.success(f"✓ נטענו {len(data)} ISINs")
            st.rerun()

st.divider()

# ── ISIN Search ───────────────────────────────────────────────────────────────
investors_db = _load_investors()

if not investors_db:
    st.info("הארכיון ריק. טען קובץ Excel למעלה.")
    st.stop()

st.markdown(f"**{len(investors_db)} ISINs בארכיון**")

col_search, col_stat = st.columns([3, 1])
with col_search:
    search_isin = st.text_input("חיפוש לפי ISIN", placeholder="XS3293111806", key="isin_search").strip().upper()
with col_stat:
    total_investors = sum(len(v["משקיעים"]) for v in investors_db.values())
    st.metric("סה\"כ ערכים", total_investors)

# ── Results ───────────────────────────────────────────────────────────────────
if search_isin:
    if search_isin in investors_db:
        prod = investors_db[search_isin]
        investors = prod["משקיעים"]

        st.markdown(f"""
<div style='background:#D6EEF2; border-radius:10px; padding:1rem 1.25rem; direction:rtl; margin-bottom:1rem;'>
  <div style='font-size:.8rem; color:#1F6B75; margin-bottom:.25rem;'>ISIN: {search_isin}</div>
  <div style='font-size:1.1rem; font-weight:700; color:#1E2761;'>{prod['שם מלא']}</div>
  <div style='font-size:.9rem; color:#555; margin-top:.25rem;'>מנפיק: {prod['מנפיק']} &nbsp;|&nbsp; תאריך הנפקה: {prod['ISSUE DATE']}</div>
</div>
""", unsafe_allow_html=True)

        if investors:
            df_inv = pd.DataFrame(investors)
            total_raised = df_inv["סכום"].sum()
            currency_counts = df_inv["מטבע"].value_counts().to_dict()

            c1, c2, c3 = st.columns(3)
            with c1:
                st.metric("מספר משקיעים", len(investors))
            with c2:
                ils_total = df_inv[df_inv["מטבע"] == "ILS"]["סכום"].sum()
                st.metric("סכום כולל (₪)", f"₪{ils_total:,.0f}")
            with c3:
                usd_total = df_inv[df_inv["מטבע"] == "USD"]["סכום"].sum()
                if usd_total:
                    st.metric("סכום כולל ($)", f"${usd_total:,.0f}")
                else:
                    st.metric("סה\"כ גויס", f"₪{total_raised:,.0f}")

            # Show table with colour by status
            def _color_status(val):
                colors = {"ACTIVE": "#C6EFCE", "CANCELLED": "#FDEBD0", "COMPLETED": "#D6EEF2"}
                return f"background-color:{colors.get(str(val).upper(), '')}"

            st.dataframe(df_inv, use_container_width=True, hide_index=True)

            # Check if any of these investors are in current Sales
            sales_df = read_df("Sales")
            if not sales_df.empty and "שם לקוח" in sales_df.columns and "ISIN פקדון" in sales_df.columns:
                current_investors = sales_df[sales_df["ISIN פקדון"] == search_isin]["שם לקוח"].tolist()
                if current_investors:
                    st.markdown(f"**📌 משקיעים מגיוס זה שנמצאים גם בגיוס הנוכחי:** {', '.join(current_investors)}")

            st.markdown("---")
            st.markdown("#### ➕ הוסף כפקדון פקוע (לידים חמים)")
            st.caption("לחץ על משקיע להוספתו לרשימת פקדונות פקועים:")

            for inv in investors:
                col_name, col_btn = st.columns([4, 1])
                with col_name:
                    st.write(f"**{inv['שם המשקיע']}** — {inv['מטבע']} {inv['סכום']:,.0f}")
                with col_btn:
                    btn_key = f"add_{search_isin}_{inv['שם המשקיע']}"
                    if st.button("הוסף →", key=btn_key):
                        row = [
                            inv["שם המשקיע"],           # שם לקוח
                            search_isin,                # ISIN
                            prod["מנפיק"],              # בנק מנפיק
                            inv["סכום"],                # סכום
                            inv.get("מטבע", "ILS"),     # מטבע
                            "",                         # תאריך פקיעה — יעודכן
                            "",                         # סכום + קופון
                            "לא",                       # פנינו לגבי חדש
                            "נמוכה",                   # רמת עניין
                            f"נטען מארכיון | {prod['שם מלא'][:50]}",  # הערות
                        ]
                        append_row("Redemptions", row)
                        log_action(current_user(), "הוסף פקדון פקוע מארכיון", f"{inv['שם המשקיע']} | {search_isin}")
                        st.success(f"✓ {inv['שם המשקיע']} נוסף לפקדונות פקועים")

            # Add all button
            st.markdown("")
            if st.button("📋 הוסף את כולם לפקדונות פקועים", use_container_width=True, type="primary"):
                added = 0
                existing_df = read_df("Redemptions")
                existing_names = existing_df["שם לקוח"].tolist() if not existing_df.empty and "שם לקוח" in existing_df.columns else []
                for inv in investors:
                    if inv["שם המשקיע"] not in existing_names:
                        row = [
                            inv["שם המשקיע"],
                            search_isin,
                            prod["מנפיק"],
                            inv["סכום"],
                            inv.get("מטבע", "ILS"),
                            "",
                            "",
                            "לא",
                            "נמוכה",
                            f"נטען מארכיון | {prod['שם מלא'][:50]}",
                        ]
                        append_row("Redemptions", row)
                        added += 1
                log_action(current_user(), "הוסף פקדונות פקועים מארכיון (כולם)", f"{search_isin} | {added} משקיעים")
                st.success(f"✓ {added} משקיעים נוספו לפקדונות פקועים")
        else:
            st.info("לא נמצאו משקיעים עבור ISIN זה.")
    else:
        # Fuzzy search — show partial matches
        matches = [isin for isin in investors_db if search_isin in isin]
        if matches:
            st.warning(f"ISIN מדויק לא נמצא. תוצאות דומות:")
            for m in matches[:10]:
                p = investors_db[m]
                st.markdown(f"- **{m}** — {p['שם מלא'][:60]} ({len(p['משקיעים'])} משקיעים)")
        else:
            st.error(f"לא נמצא ISIN: {search_isin}")

# ── Full Index ────────────────────────────────────────────────────────────────
st.divider()
with st.expander("📋 כל הISINs בארכיון", expanded=False):
    index_rows = []
    for isin, data in investors_db.items():
        total_amt = sum(i["סכום"] for i in data["משקיעים"])
        index_rows.append({
            "ISIN": isin,
            "שם מוצר": data["שם מלא"][:60],
            "מנפיק": data["מנפיק"],
            "תאריך הנפקה": data["ISSUE DATE"],
            "מספר משקיעים": len(data["משקיעים"]),
            "סכום כולל": f'₪{total_amt:,.0f}',
        })
    st.dataframe(pd.DataFrame(index_rows), use_container_width=True, hide_index=True)
